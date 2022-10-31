import openpyxl as opx
from openpyxl.styles import PatternFill,Alignment,Border,Font,Side

class XlsxTool():
    def __init__(self):
        self.wb = None
        self.sheetNum = 0
        self.sheet = None
        self.path = None
        self.fontlib = self.__loadFontLib__()
        self.colorLib = {'red':'FF0000','blue':'0000FF','green':'00FF00','black':'000000','white':'FFFFFF','yellow':'FFFF00','gray':'808080','silver':'C0C0C0'}
        self.borderLib = self.__loadBorderLib__()

    def read(self,path):
        if self.__check__(path) == 1:
            self.wb = opx.load_workbook(path)
            self.sheetNum = len(self.wb.sheetnames)
            self.path = path
        else:
            raise Exception('invalidPath')

    def selectSheet(self,sheetName):
        if self.__check__(sheetName,list=self.wb.sheetnames) == 1:
            self.sheet = self.wb[sheetName]
        else:
            raise Exception('invalidPath')

    def save(self,path=None):
        if path is None:
            self.wb.save(self.path)
        else:
            if self.wb is None:
                self.wb = opx.Workbook()
                self.wb.save(path)
            else:
                self.wb.save(path)

    def applyCells(self,start,end,filter):
        if end is None:
            end  = start
        range = self.sheet[start:end]
        for row in range:
            for cell in row:
                filter(cell)

    def decorate(self,start,end=None,fill=None,border=None,font=None,align=None):
        if end is None and start is not None:
            end = start
        elif end is None and start is None:
            print('range not exist')
            return None
        range = self.sheet[start:end]
        for row in range:
            for cell in row:
                if fill is not None:
                    cell.fill = PatternFill(patternType='solid',fgColor=fill)
                if border is not None:
                    cell.border = Border(left=border,right=border,bottom=border,top=border)
                if font is not None:
                    cell.font = font
                if align is not None:
                    cell.alignment = Alignment(horizontal=align,vertical='center',wrapText=True)
        return True

    def resize(self,rows=None,height=None,cols=None,width=None):
        if rows is not None and height is not None:
            for row in rows:
                self.sheet.row_dimensions[row].height = height
        if cols is not None and width is not None:
            for col in cols:
                self.sheet.column_dimensions[col].width = width

    def freeze(self,range):
        self.freeze_panes = range

    def merge(self,start,end):
        if start is None:
            return None
        if end is None:
            end = start
        range = start+':'+end
        self.sheet.merge_cells(range)
        self.decorate(start=start,align='center')
        return True

    def __rgb2str__(self,channels):
        color = ''
        for channel in channels:
            if channel > 255 or channel < 0:
                print('only 256 bit RGB is supported')
                return None
            temp=hex(channel)[2:]
            if len(temp) == 1:
                temp = '0'+temp
            color+=temp
        return color

    def getColor(self,color):
        if color in self.colorLib.keys():
            return self.colorLib[color]
        else:
            print('color not defined')
            return None

    def getFont(self,font):
        if font in self.fontlib.keys():
            return self.fontlib[font]
        else:
            print('font is not defined')
            return None

    def getBorder(self,side):
        if side in self.borderLib.keys():
            return self.borderLib[side]
        else:
            print('border is not defined')
            return None

    def __strIndex2Int__(self, index):
        colNum = 0
        for i in list(range(0 - len(index), 0)):
            colNum += (ord(index[i]) - 64) * pow(26, (-(i + 1)))
        colNum -= 1
        return colNum

    def __check__(self,val,min=None,max=None,list=None):
        if val is None:
            return 0
        if isinstance(val,int):
            if min is not None and max is not None:
                if val >= min and val <= max:
                    return 1
                else:
                    return 2
        elif isinstance(val,str):
            if list is not None:
                if val in list:
                    return 1
                else:
                    return 3
            else:
                return 1
        return None

    def __loadFontLib__(self):
        fontlib = {}
        fontlib['PATAC'] = Font(name='Arial',size='10',color='000000')
        fontlib['PATAC_Bold'] = Font(name='Arial',size='10',bold=True,color='000000')
        fontlib['PATAC_Italic'] = Font(name='Arial',size='10',italic=True,color='000000')
        return fontlib

    def __loadBorderLib__(self):
        borderLib = {}
        borderLib['normal']=Side(style='thin',color='000000')
        borderLib['important'] = Side(style='double',color='FF0000')
        borderLib['normalBold'] = Side(style='thick',color='000000')
        borderLib['dottedBlue'] = Side(style='dotted',color='0000FF')
        return borderLib

if __name__ == '__main__':
    xt = XlsxTool()
    path = r'../testdata/travel-budget.xlsx'
    xt.read(path)
    print(xt.sheetNum)
    xt.selectSheet("test2")
    print(xt.sheet['A1'].value)
    xt.sheet['A1']=3
    xt.decorate(start='A1',end='A4',fill=xt.getColor('red'),border=xt.getBorder('normal'),align='center',font=xt.getFont('PATAC_Bold'))
    xt.merge('F9','G9')
    xt.resize(rows=[1],height=16,cols=['A'],width=10)
    xt.selectSheet('Travel Budget')
    xt.sheet['B1'].value = 0
    path = r'../testdata/test2.xlsx'
    xt.save(path)
